from __future__ import annotations

import json
import sys
import uuid
from pathlib import Path
from typing import TYPE_CHECKING, Any

import click
from rich.console import Console
from rich.panel import Panel
from rich.table import Table

from opennotes_cli.display import get_cli_prefix, get_status_style, handle_jsonapi_error
from opennotes_cli.formatting import format_id, resolve_id
from opennotes_cli.http import add_csrf, get_csrf_token
from opennotes_cli.polling import poll_batch_job_until_complete, poll_simulation_until_complete

if TYPE_CHECKING:
    from opennotes_cli.cli import CliContext

console = Console()
error_console = Console(stderr=True)


@click.group()
def simulation() -> None:
    """Manage simulation runs."""


@simulation.command("list")
@click.option("--page", default=1, type=click.IntRange(1), help="Page number.")
@click.option("--page-size", default=20, type=click.IntRange(1, 100), help="Page size (1-100).")
@click.pass_context
def simulation_list(ctx: click.Context, page: int, page_size: int) -> None:
    """List simulation runs."""
    cli_ctx: CliContext = ctx.obj
    base_url = cli_ctx.base_url
    client = cli_ctx.client

    csrf_token = get_csrf_token(client, base_url, cli_ctx.auth)
    headers = add_csrf(cli_ctx.auth.get_jsonapi_headers(), csrf_token)

    params = {"page[number]": page, "page[size]": page_size}
    response = client.get(f"{base_url}/api/v2/simulations", headers=headers, params=params)
    handle_jsonapi_error(response)

    result = response.json()

    if cli_ctx.json_output:
        console.print(json.dumps(result, indent=2, default=str))
        return

    items = result.get("data", [])
    total = result.get("meta", {}).get("count", len(items))

    if not items:
        console.print("[yellow]No simulation runs found.[/yellow]")
        return

    table = Table(show_header=True, header_style="bold")
    table.add_column("ID")
    table.add_column("Status", width=12)
    table.add_column("Orchestrator ID")
    table.add_column("Community Server ID")
    table.add_column("Created At", width=20)

    for item in items:
        attrs = item.get("attributes", {})
        status = attrs.get("status", "unknown")
        color, symbol = get_status_style(status)
        created = (attrs.get("created_at") or "")[:19]

        table.add_row(
            format_id(item.get("id", "N/A"), cli_ctx.use_huuid),
            f"[{color}]{symbol} {status}[/{color}]",
            format_id(attrs.get("orchestrator_id", "N/A"), cli_ctx.use_huuid),
            format_id(attrs.get("community_server_id", "N/A"), cli_ctx.use_huuid),
            created,
        )

    console.print(table)
    console.print(f"\n[dim]Total: {total} simulation(s)[/dim]")


@simulation.command("status")
@click.argument("simulation_id")
@click.pass_context
def simulation_status(ctx: click.Context, simulation_id: str) -> None:
    """Show status of a simulation run."""
    try:
        simulation_id = resolve_id(simulation_id)
    except click.BadParameter as e:
        error_console.print(f"[red]Error:[/red] {e}")
        sys.exit(1)

    cli_ctx: CliContext = ctx.obj
    base_url = cli_ctx.base_url
    client = cli_ctx.client

    csrf_token = get_csrf_token(client, base_url, cli_ctx.auth)
    headers = add_csrf(cli_ctx.auth.get_jsonapi_headers(), csrf_token)

    response = client.get(
        f"{base_url}/api/v2/simulations/{simulation_id}", headers=headers
    )
    handle_jsonapi_error(response)

    result = response.json()

    if cli_ctx.json_output:
        console.print(json.dumps(result, indent=2, default=str))
        return

    attrs = result.get("data", {}).get("attributes", {})
    status = attrs.get("status", "unknown")
    color, symbol = get_status_style(status)

    panel_content = (
        f"[bold]ID:[/bold] {format_id(result.get('data', {}).get('id', 'N/A'), cli_ctx.use_huuid)}\n"
        f"[bold]Status:[/bold] [{color}]{symbol} {status}[/{color}]\n"
        f"[bold]Orchestrator:[/bold] {format_id(attrs.get('orchestrator_id', 'N/A'), cli_ctx.use_huuid)}\n"
        f"[bold]Community Server:[/bold] {format_id(attrs.get('community_server_id', 'N/A'), cli_ctx.use_huuid)}"
    )

    for ts_field in ("started_at", "completed_at", "paused_at", "created_at", "updated_at"):
        val = attrs.get(ts_field)
        if val:
            label = ts_field.replace("_", " ").title()
            panel_content += f"\n[bold]{label}:[/bold] {val[:19]}"

    metrics = attrs.get("metrics")
    if metrics:
        panel_content += f"\n[bold]Metrics:[/bold] {json.dumps(metrics, indent=2)}"

    restart_count = attrs.get("restart_count", 0)
    if restart_count:
        panel_content += f"\n[bold]Restarts:[/bold] {restart_count}"
        cumulative = attrs.get("cumulative_turns", 0)
        panel_content += f"\n[bold]Cumulative Turns:[/bold] {cumulative}"

    error_msg = attrs.get("error_message")
    if error_msg:
        panel_content += f"\n[bold red]Error:[/bold red] {error_msg}"

    try:
        progress_response = client.get(
            f"{base_url}/api/v2/simulations/{simulation_id}/progress",
            headers=headers,
        )
        if progress_response.status_code == 200:
            progress = progress_response.json().get("data", {}).get("attributes", {})
            panel_content += "\n\n[bold]Progress:[/bold]"
            turns_completed = progress.get("turns_completed", 0)
            turns_errored = progress.get("turns_errored", 0)
            panel_content += f"\n  Turns: {turns_completed} completed"
            if turns_errored:
                panel_content += f", {turns_errored} errored"
            panel_content += f"\n  Notes: {progress.get('notes_written', 0)}"
            panel_content += f"\n  Ratings: {progress.get('ratings_given', 0)}"
            panel_content += f"\n  Active Agents: {progress.get('active_agents', 0)}"
    except Exception:
        pass

    console.print(Panel(panel_content, title="[bold]Simulation Run[/bold]"))


@simulation.command("create")
@click.option("--orchestrator-id", required=True, help="Orchestrator UUID.")
@click.option(
    "--community-server-id", required=True,
    help="Community server UUID (must be playground).",
)
@click.option("--wait", is_flag=True, help="Poll until simulation reaches a terminal state.")
@click.option("--copy-requests-from", default=None, help="Copy requests from this community server ID before creating simulation.")
@click.pass_context
def simulation_create(
    ctx: click.Context, orchestrator_id: str, community_server_id: str, wait: bool, copy_requests_from: str | None
) -> None:
    """Create and start a simulation run."""
    try:
        orchestrator_id = resolve_id(orchestrator_id)
    except click.BadParameter as e:
        error_console.print(f"[red]Error:[/red] {e}")
        sys.exit(1)
    try:
        community_server_id = resolve_id(community_server_id)
    except click.BadParameter as e:
        error_console.print(f"[red]Error:[/red] {e}")
        sys.exit(1)

    cli_ctx: CliContext = ctx.obj
    base_url = cli_ctx.base_url
    client = cli_ctx.client

    csrf_token = get_csrf_token(client, base_url, cli_ctx.auth)
    headers = add_csrf(cli_ctx.auth.get_jsonapi_headers(), csrf_token)

    if copy_requests_from:
        try:
            copy_source_id = resolve_id(copy_requests_from)
        except click.BadParameter as e:
            error_console.print(f"[red]Error:[/red] {e}")
            sys.exit(1)

        copy_payload = {
            "data": {
                "type": "copy-requests",
                "attributes": {
                    "source_community_server_id": copy_source_id,
                },
            }
        }
        copy_response = client.post(
            f"{base_url}/api/v2/community-servers/{community_server_id}/copy-requests",
            headers=headers,
            json=copy_payload,
        )
        handle_jsonapi_error(copy_response)
        copy_job_id = copy_response.json().get("data", {}).get("id")

        if not cli_ctx.json_output:
            console.print(f"[green]\u2713[/green] Copying requests from {format_id(copy_source_id, cli_ctx.use_huuid)}...")

        auth_headers = add_csrf(cli_ctx.auth.get_headers(), csrf_token)
        poll_batch_job_until_complete(client, base_url, auth_headers, copy_job_id, use_huuid=cli_ctx.use_huuid)

        if not cli_ctx.json_output:
            console.print("[green]\u2713[/green] Requests copied successfully")

    payload = {
        "data": {
            "type": "simulations",
            "attributes": {
                "orchestrator_id": orchestrator_id,
                "community_server_id": community_server_id,
            },
        }
    }

    response = client.post(
        f"{base_url}/api/v2/simulations", headers=headers, json=payload
    )
    handle_jsonapi_error(response)

    result = response.json()
    sim_id = result.get("data", {}).get("id", "N/A")

    if wait:
        if not cli_ctx.json_output:
            console.print(f"[dim]Simulation created: {format_id(sim_id, cli_ctx.use_huuid)}[/dim]")
            console.print("[dim]Waiting for completion...[/dim]\n")
        result = poll_simulation_until_complete(client, base_url, headers, sim_id, use_huuid=cli_ctx.use_huuid)

    if cli_ctx.json_output:
        console.print(json.dumps(result, indent=2, default=str))
        return

    attrs = result.get("data", {}).get("attributes", {})
    status = attrs.get("status", "unknown")
    color, symbol = get_status_style(status)

    console.print(
        Panel(
            f"[bold]ID:[/bold] {format_id(sim_id, cli_ctx.use_huuid)}\n"
            f"[bold]Status:[/bold] [{color}]{symbol} {status}[/{color}]\n"
            f"[bold]Orchestrator:[/bold] {format_id(attrs.get('orchestrator_id', 'N/A'), cli_ctx.use_huuid)}\n"
            f"[bold]Community Server:[/bold] {format_id(attrs.get('community_server_id', 'N/A'), cli_ctx.use_huuid)}",
            title="[bold green]Simulation Created[/bold green]",
        )
    )

    if not wait:
        cli_prefix = get_cli_prefix(cli_ctx.env_name)
        console.print(f"\n[dim]Check status:[/dim] {cli_prefix} simulation status {format_id(sim_id, cli_ctx.use_huuid)}")


def _simulation_lifecycle_cmd(
    ctx: click.Context, simulation_id: str, action: str
) -> None:
    cli_ctx: CliContext = ctx.obj
    base_url = cli_ctx.base_url
    client = cli_ctx.client

    csrf_token = get_csrf_token(client, base_url, cli_ctx.auth)
    headers = add_csrf(cli_ctx.auth.get_jsonapi_headers(), csrf_token)

    response = client.post(
        f"{base_url}/api/v2/simulations/{simulation_id}/{action}", headers=headers
    )
    handle_jsonapi_error(response)

    result = response.json()

    if cli_ctx.json_output:
        console.print(json.dumps(result, indent=2, default=str))
        return

    attrs = result.get("data", {}).get("attributes", {})
    status = attrs.get("status", "unknown")
    color, symbol = get_status_style(status)
    console.print(
        f"[{color}]{symbol}[/{color}] Simulation {format_id(simulation_id, cli_ctx.use_huuid)} is now [{color}]{status}[/{color}]"
    )


@simulation.command("pause")
@click.argument("simulation_id")
@click.pass_context
def simulation_pause(ctx: click.Context, simulation_id: str) -> None:
    """Pause a running simulation."""
    try:
        simulation_id = resolve_id(simulation_id)
    except click.BadParameter as e:
        error_console.print(f"[red]Error:[/red] {e}")
        sys.exit(1)
    _simulation_lifecycle_cmd(ctx, simulation_id, "pause")


@simulation.command("resume")
@click.argument("simulation_id")
@click.option(
    "--reset-turns",
    is_flag=True,
    default=False,
    help="Reset turn counts (required for cancelled runs).",
)
@click.pass_context
def simulation_resume(ctx: click.Context, simulation_id: str, reset_turns: bool) -> None:
    """Resume a paused, pending, or cancelled simulation."""
    try:
        simulation_id = resolve_id(simulation_id)
    except click.BadParameter as e:
        error_console.print(f"[red]Error:[/red] {e}")
        sys.exit(1)

    if reset_turns:
        cli_ctx: CliContext = ctx.obj
        base_url = cli_ctx.base_url
        client = cli_ctx.client

        csrf_token = get_csrf_token(client, base_url, cli_ctx.auth)
        headers = add_csrf(cli_ctx.auth.get_jsonapi_headers(), csrf_token)

        body = {
            "data": {
                "type": "simulations",
                "attributes": {"reset_turns": True},
            }
        }

        response = client.post(
            f"{base_url}/api/v2/simulations/{simulation_id}/resume",
            headers=headers,
            json=body,
        )
        handle_jsonapi_error(response)

        result = response.json()

        if cli_ctx.json_output:
            console.print(json.dumps(result, indent=2, default=str))
            return

        attrs = result.get("data", {}).get("attributes", {})
        status_val = attrs.get("status", "unknown")
        color, symbol = get_status_style(status_val)
        console.print(
            f"[{color}]{symbol}[/{color}] Simulation {format_id(simulation_id, cli_ctx.use_huuid)} resumed with turn reset, "
            f"now [{color}]{status_val}[/{color}]"
        )
    else:
        _simulation_lifecycle_cmd(ctx, simulation_id, "resume")


@simulation.command("cancel")
@click.argument("simulation_id")
@click.pass_context
def simulation_cancel(ctx: click.Context, simulation_id: str) -> None:
    """Cancel a simulation."""
    try:
        simulation_id = resolve_id(simulation_id)
    except click.BadParameter as e:
        error_console.print(f"[red]Error:[/red] {e}")
        sys.exit(1)
    _simulation_lifecycle_cmd(ctx, simulation_id, "cancel")


@simulation.command("cancel-workflows")
@click.argument("simulation_id")
@click.option("--generation", type=int, default=None, help="Only cancel workflows from this generation.")
@click.option("--dry-run", is_flag=True, default=False, help="List workflows without cancelling.")
@click.pass_context
def simulation_cancel_workflows(
    ctx: click.Context, simulation_id: str, generation: int | None, dry_run: bool
) -> None:
    """List and bulk-cancel orphaned turn workflows for a simulation."""
    try:
        simulation_id = resolve_id(simulation_id)
    except click.BadParameter as e:
        error_console.print(f"[red]Error:[/red] {e}")
        sys.exit(1)

    cli_ctx: CliContext = ctx.obj
    base_url = cli_ctx.base_url
    client = cli_ctx.client

    csrf_token = get_csrf_token(client, base_url, cli_ctx.auth)
    headers = add_csrf(cli_ctx.auth.get_jsonapi_headers(), csrf_token)

    params: dict[str, str | int] = {}
    if dry_run:
        params["dry_run"] = "true"
    if generation is not None:
        params["generation"] = generation

    response = client.post(
        f"{base_url}/api/v2/simulations/{simulation_id}/cancel-workflows",
        headers=headers,
        params=params,
    )
    handle_jsonapi_error(response)

    result = response.json()

    if cli_ctx.json_output:
        console.print(json.dumps(result, indent=2, default=str))
        return

    workflow_ids = result.get("workflow_ids", [])
    total = result.get("total", 0)
    cancelled = result.get("cancelled", 0)

    if dry_run:
        console.print(f"[yellow]Dry run:[/yellow] found {total} workflow(s)")
    else:
        console.print(f"[green]Cancelled {cancelled}[/green] of {total} workflow(s)")

    if workflow_ids:
        table = Table(show_header=True, header_style="bold")
        table.add_column("Workflow ID")
        for wf_id in workflow_ids:
            table.add_row(format_id(wf_id, cli_ctx.use_huuid))
        console.print(table)
    else:
        console.print("[dim]No workflows found.[/dim]")


@simulation.command("restart")
@click.argument("simulation_id")
@click.pass_context
def simulation_restart(ctx: click.Context, simulation_id: str) -> None:
    """Restart a completed simulation (reset turns, keep memory)."""
    try:
        simulation_id = resolve_id(simulation_id)
    except click.BadParameter as e:
        error_console.print(f"[red]Error:[/red] {e}")
        sys.exit(1)

    cli_ctx: CliContext = ctx.obj
    base_url = cli_ctx.base_url
    client = cli_ctx.client

    csrf_token = get_csrf_token(client, base_url, cli_ctx.auth)
    headers = add_csrf(cli_ctx.auth.get_jsonapi_headers(), csrf_token)

    body = {
        "data": {
            "type": "simulations",
            "attributes": {"reset_turns": True},
        }
    }

    response = client.post(
        f"{base_url}/api/v2/simulations/{simulation_id}/resume",
        headers=headers,
        json=body,
    )
    handle_jsonapi_error(response)

    result = response.json()

    if cli_ctx.json_output:
        console.print(json.dumps(result, indent=2, default=str))
        return

    attrs = result.get("data", {}).get("attributes", {})
    status_val = attrs.get("status", "unknown")
    color, symbol = get_status_style(status_val)
    restart_count = attrs.get("restart_count", 0)
    console.print(
        f"[{color}]{symbol}[/{color}] Simulation {format_id(simulation_id, cli_ctx.use_huuid)} restarted "
        f"(restart #{restart_count}), now [{color}]{status_val}[/{color}]"
    )


def _read_urls_from_file(path: str) -> list[str]:
    filepath = Path(path)
    if not filepath.exists():
        error_console.print(f"[red]Error:[/red] File not found: {filepath}")
        sys.exit(1)
    urls: list[str] = []
    for line in filepath.read_text().splitlines():
        line = line.strip()
        if line.startswith(("- ", "* ")):
            line = line[2:].strip()
        if line and line.startswith("http"):
            urls.append(line)
    return urls


def _submit_urls_batch(
    client: Any,
    base_url: str,
    headers: dict[str, str],
    community_server_id: str,
    urls: list[str],
) -> str | None:
    payload = {
        "data": {
            "type": "playground-note-requests",
            "attributes": {"urls": urls},
        }
    }
    response = client.post(
        f"{base_url}/api/v2/playgrounds/{community_server_id}/note-requests",
        headers=headers,
        json=payload,
    )
    handle_jsonapi_error(response)
    return response.json().get("data", {}).get("attributes", {}).get("workflow_id")


@simulation.command("launch")
@click.option("--name", default=None, help="Simulation name (used for orchestrator and playground).")
@click.option("--url", "urls", multiple=True, help="URL to add as note request (repeatable).")
@click.option("--urls-from", "urls_file", default=None, help="File with URLs, one per line.")
@click.option("--text", "texts", multiple=True, help="Text content to submit directly (repeatable).")
@click.option("--agent-ids", default=None, help="Comma-separated agent profile UUIDs (default: all).")
@click.option("--max-agents", default=15, type=int, help="Max concurrent agents (default: 15).")
@click.option("--turn-cadence", default=15, type=int, help="Turn cadence in seconds (default: 15).")
@click.option("--removal-rate", default=0.0, type=float, help="Agent removal rate (default: 0.0).")
@click.option("--max-turns", default=50, type=int, help="Max turns per agent (default: 50).")
@click.option("--wait", is_flag=True, help="Poll until simulation completes.")
@click.option("--copy-requests-from", default=None, help="Copy requests from this community server ID before running simulation.")
@click.pass_context
def simulation_launch(
    ctx: click.Context,
    name: str | None,
    urls: tuple[str, ...],
    urls_file: str | None,
    texts: tuple[str, ...],
    agent_ids: str | None,
    max_agents: int,
    turn_cadence: int,
    removal_rate: float,
    max_turns: int,
    wait: bool,
    copy_requests_from: str | None,
) -> None:
    """Launch a full simulation: create orchestrator, playground, submit content, start sim."""
    cli_ctx: CliContext = ctx.obj
    base_url = cli_ctx.base_url
    client = cli_ctx.client

    all_urls = list(urls)
    if urls_file:
        all_urls.extend(_read_urls_from_file(urls_file))

    if not all_urls and not texts:
        error_console.print("[red]Error:[/red] Provide at least one --url, --urls-from, or --text.")
        sys.exit(1)

    csrf_token = get_csrf_token(client, base_url, cli_ctx.auth)
    headers = add_csrf(cli_ctx.auth.get_jsonapi_headers(), csrf_token)

    if not name:
        name = f"sim-{uuid.uuid4().hex[:8]}"

    if agent_ids:
        agent_id_list = [a.strip() for a in agent_ids.split(",") if a.strip()]
        resolved_agent_ids: list[str] = []
        for aid in agent_id_list:
            try:
                resolved_agent_ids.append(resolve_id(aid))
            except click.BadParameter as e:
                error_console.print(f"[red]Error:[/red] {e}")
                sys.exit(1)
        agent_id_list = resolved_agent_ids
    else:
        response = client.get(
            f"{base_url}/api/v2/sim-agents",
            headers=headers,
            params={"page[size]": 100},
        )
        handle_jsonapi_error(response)
        items = response.json().get("data", [])
        agent_id_list = [item["id"] for item in items]
        if not agent_id_list:
            error_console.print("[red]Error:[/red] No sim agents found. Create agents first.")
            sys.exit(1)
        if not cli_ctx.json_output:
            console.print(f"[dim]Discovered {len(agent_id_list)} agent(s)[/dim]")

    orch_attributes: dict[str, Any] = {
        "name": name,
        "turn_cadence_seconds": turn_cadence,
        "max_agents": max_agents,
        "removal_rate": removal_rate,
        "max_turns_per_agent": max_turns,
        "agent_profile_ids": agent_id_list,
    }
    orch_payload = {"data": {"type": "simulation-orchestrators", "attributes": orch_attributes}}
    response = client.post(
        f"{base_url}/api/v2/simulation-orchestrators", headers=headers, json=orch_payload
    )
    handle_jsonapi_error(response)
    orch_id = response.json().get("data", {}).get("id")
    if not cli_ctx.json_output:
        console.print(f"[green]\u2713[/green] Created orchestrator: [bold]{format_id(orch_id, cli_ctx.use_huuid)}[/bold]")

    pg_headers = add_csrf(cli_ctx.auth.get_headers(), csrf_token)
    pg_payload: dict[str, Any] = {
        "platform": "playground",
        "platform_community_server_id": uuid.uuid4().hex[:16],
        "name": name,
        "is_active": True,
        "is_public": True,
    }
    response = client.post(
        f"{base_url}/api/v1/community-servers", headers=pg_headers, json=pg_payload
    )
    handle_jsonapi_error(response)
    cs_id = response.json().get("id")
    if not cli_ctx.json_output:
        console.print(f"[green]\u2713[/green] Created playground: [bold]{format_id(cs_id, cli_ctx.use_huuid)}[/bold]")

    jsonapi_headers = add_csrf(cli_ctx.auth.get_jsonapi_headers(), csrf_token)

    if copy_requests_from:
        try:
            copy_source_id = resolve_id(copy_requests_from)
        except click.BadParameter as e:
            error_console.print(f"[red]Error:[/red] {e}")
            sys.exit(1)

        copy_payload = {
            "data": {
                "type": "copy-requests",
                "attributes": {
                    "source_community_server_id": copy_source_id,
                },
            }
        }
        copy_response = client.post(
            f"{base_url}/api/v2/community-servers/{cs_id}/copy-requests",
            headers=jsonapi_headers,
            json=copy_payload,
        )
        handle_jsonapi_error(copy_response)
        copy_job_id = copy_response.json().get("data", {}).get("id")

        if not cli_ctx.json_output:
            console.print(f"[green]\u2713[/green] Copying requests from {format_id(copy_source_id, cli_ctx.use_huuid)}...")

        auth_headers = add_csrf(cli_ctx.auth.get_headers(), csrf_token)
        poll_batch_job_until_complete(client, base_url, auth_headers, copy_job_id, use_huuid=cli_ctx.use_huuid)

        if not cli_ctx.json_output:
            console.print("[green]\u2713[/green] Requests copied successfully")

    if all_urls:
        for i in range(0, len(all_urls), 20):
            batch = all_urls[i : i + 20]
            _submit_urls_batch(client, base_url, jsonapi_headers, cs_id, batch)
        if not cli_ctx.json_output:
            console.print(f"[green]\u2713[/green] Submitted {len(all_urls)} URL(s) for processing")

    if texts:
        all_texts = list(texts)
        for i in range(0, len(all_texts), 20):
            batch = all_texts[i : i + 20]
            text_payload = {
                "data": {
                    "type": "playground-note-requests",
                    "attributes": {"texts": batch},
                }
            }
            response = client.post(
                f"{base_url}/api/v2/playgrounds/{cs_id}/note-requests",
                headers=jsonapi_headers,
                json=text_payload,
            )
            handle_jsonapi_error(response)
        if not cli_ctx.json_output:
            console.print(f"[green]\u2713[/green] Submitted {len(texts)} text(s) for processing")

    sim_payload = {
        "data": {
            "type": "simulations",
            "attributes": {
                "orchestrator_id": orch_id,
                "community_server_id": cs_id,
            },
        }
    }
    response = client.post(
        f"{base_url}/api/v2/simulations", headers=jsonapi_headers, json=sim_payload
    )
    handle_jsonapi_error(response)
    result = response.json()
    sim_id = result.get("data", {}).get("id")

    if not cli_ctx.json_output:
        console.print(f"[green]\u2713[/green] Started simulation: [bold]{format_id(sim_id, cli_ctx.use_huuid)}[/bold]")

    if wait:
        if not cli_ctx.json_output:
            console.print("[dim]Waiting for completion...[/dim]\n")
        result = poll_simulation_until_complete(client, base_url, jsonapi_headers, sim_id, use_huuid=cli_ctx.use_huuid)

    if cli_ctx.json_output:
        console.print(json.dumps({
            "orchestrator_id": orch_id,
            "community_server_id": cs_id,
            "simulation_id": sim_id,
            "result": result,
        }, indent=2, default=str))
        return

    cli_prefix = get_cli_prefix(cli_ctx.env_name)
    console.print(
        f"\n[dim]Monitor:[/dim]  {cli_prefix} simulation status {format_id(sim_id, cli_ctx.use_huuid)}"
    )
    console.print(
        f"[dim]Update:[/dim]   {cli_prefix} orchestrator apply {format_id(orch_id, cli_ctx.use_huuid)} --simulation {format_id(sim_id, cli_ctx.use_huuid)} --max-agents 20"
    )


@simulation.command("analysis")
@click.argument("simulation_id")
@click.option(
    "--format",
    "output_format",
    type=click.Choice(["terminal", "markdown", "xlsx"]),
    default="terminal",
    help="Output format.",
)
@click.option("--detailed", is_flag=True, help="Fetch detailed per-note/rating/request data.")
@click.option("--output", "output_path", default=None, help="Output file path (for xlsx format).")
@click.pass_context
def simulation_analysis(
    ctx: click.Context, simulation_id: str, output_format: str, detailed: bool, output_path: str | None
) -> None:
    """Show analysis results for a simulation run."""
    try:
        simulation_id = resolve_id(simulation_id)
    except click.BadParameter as e:
        error_console.print(f"[red]Error:[/red] {e}")
        sys.exit(1)

    cli_ctx: CliContext = ctx.obj
    base_url = cli_ctx.base_url
    client = cli_ctx.client
    use_huuid = cli_ctx.use_huuid

    if output_format == "xlsx" and not detailed:
        error_console.print("[red]Error:[/red] --format xlsx requires --detailed flag.")
        sys.exit(1)

    csrf_token = get_csrf_token(client, base_url, cli_ctx.auth)
    headers = add_csrf(cli_ctx.auth.get_jsonapi_headers(), csrf_token)

    if detailed:
        accumulated = _fetch_detailed_pages(client, base_url, headers, simulation_id)

        if cli_ctx.json_output:
            console.print(json.dumps(accumulated, indent=2, default=str))
            return

        if output_format == "markdown":
            _render_detailed_markdown(simulation_id, accumulated, use_huuid)
        elif output_format == "xlsx":
            _render_detailed_xlsx(simulation_id, accumulated, output_path, use_huuid)
        else:
            _render_detailed_terminal(simulation_id, accumulated, use_huuid)
        return

    response = client.get(
        f"{base_url}/api/v2/simulations/{simulation_id}/analysis", headers=headers
    )
    handle_jsonapi_error(response)

    result = response.json()

    if cli_ctx.json_output:
        console.print(json.dumps(result, indent=2, default=str))
        return

    attrs = result.get("data", {}).get("attributes", {})

    if _is_empty_analysis(attrs):
        if output_format == "markdown":
            click.echo(f"# Simulation Analysis: {format_id(simulation_id, use_huuid)}\n\nAnalysis not available yet — simulation may still be running.")
        else:
            console.print("[yellow]Analysis not available yet — simulation may still be running.[/yellow]")
        return

    if output_format == "markdown":
        _render_analysis_markdown(simulation_id, attrs, use_huuid)
    else:
        _render_analysis_terminal(attrs)


def _is_empty_analysis(attrs: dict[str, Any]) -> bool:
    if not attrs:
        return True
    total_ratings = attrs.get("rating_distribution", {}).get("total_ratings", 0)
    total_notes = attrs.get("consensus_metrics", {}).get("total_notes_rated", 0)
    total_scores = attrs.get("scoring_coverage", {}).get("total_scores_computed", 0)
    return total_ratings == 0 and total_notes == 0 and total_scores == 0


def _render_analysis_terminal(attrs: dict[str, Any]) -> None:
    rating_dist = attrs.get("rating_distribution", {})
    overall = rating_dist.get("overall", {})
    total_ratings = rating_dist.get("total_ratings", 0)

    table = Table(title="Rating Distribution", show_header=True, header_style="bold")
    table.add_column("Rating")
    table.add_column("Count", justify="right")
    table.add_column("Percentage", justify="right")
    for rating, count in sorted(overall.items()):
        pct = (count / total_ratings * 100) if total_ratings > 0 else 0
        table.add_row(rating, str(count), f"{pct:.1f}%")
    total_pct = "100.0%" if total_ratings > 0 else "N/A"
    table.add_row("[bold]Total[/bold]", f"[bold]{total_ratings}[/bold]", f"[bold]{total_pct}[/bold]")
    console.print(table)

    per_agent = rating_dist.get("per_agent", [])
    if per_agent:
        agent_table = Table(
            title="Per-Agent Rating Breakdown", show_header=True, header_style="bold"
        )
        agent_table.add_column("Agent")
        agent_table.add_column("Ratings", justify="right")
        agent_table.add_column("Distribution")
        for agent in per_agent:
            dist_parts = [f"{k}: {v}" for k, v in agent.get("distribution", {}).items()]
            agent_table.add_row(
                agent.get("agent_name", "N/A"),
                str(agent.get("total", 0)),
                ", ".join(dist_parts),
            )
        console.print(agent_table)

    consensus = attrs.get("consensus_metrics", {})
    consensus_content = (
        f"[bold]Mean agreement:[/bold] {consensus.get('mean_agreement', 'N/A')}\n"
        f"[bold]Polarization index:[/bold] {consensus.get('polarization_index', 'N/A')}\n"
        f"[bold]Notes with consensus:[/bold] {consensus.get('notes_with_consensus', 'N/A')}\n"
        f"[bold]Notes with disagreement:[/bold] {consensus.get('notes_with_disagreement', 'N/A')}\n"
        f"[bold]Total notes rated:[/bold] {consensus.get('total_notes_rated', 'N/A')}"
    )
    console.print(Panel(consensus_content, title="[bold]Consensus Metrics[/bold]"))

    scoring = attrs.get("scoring_coverage", {})
    scoring_content = (
        f"[bold]Current tier:[/bold] {scoring.get('current_tier', 'N/A')}\n"
        f"[bold]Total scores computed:[/bold] {scoring.get('total_scores_computed', 'N/A')}"
    )
    console.print(Panel(scoring_content, title="[bold]Scoring Coverage[/bold]"))

    tier_dist = scoring.get("tier_distribution", {})
    if tier_dist:
        tier_table = Table(show_header=True, header_style="bold")
        tier_table.add_column("Tier")
        tier_table.add_column("Count", justify="right")
        for tier, count in sorted(tier_dist.items()):
            tier_table.add_row(tier, str(count))
        console.print(tier_table)

    scorer_bd = scoring.get("scorer_breakdown", {})
    if scorer_bd:
        scorer_table = Table(show_header=True, header_style="bold")
        scorer_table.add_column("Scorer")
        scorer_table.add_column("Count", justify="right")
        for scorer, count in sorted(scorer_bd.items()):
            scorer_table.add_row(scorer, str(count))
        console.print(scorer_table)

    scoring_notes = scoring.get("notes_by_status", {})
    if scoring_notes:
        sn_table = Table(show_header=True, header_style="bold")
        sn_table.add_column("Status")
        sn_table.add_column("Count", justify="right")
        for status, count in sorted(scoring_notes.items()):
            sn_table.add_row(status, str(count))
        console.print(sn_table)

    behaviors = attrs.get("agent_behaviors", [])
    if behaviors:
        beh_table = Table(title="Agent Behaviors", show_header=True, header_style="bold")
        beh_table.add_column("Agent")
        beh_table.add_column("Notes", justify="right")
        beh_table.add_column("Ratings", justify="right")
        beh_table.add_column("Turns", justify="right")
        beh_table.add_column("State")
        beh_table.add_column("Top Actions")
        for agent in behaviors:
            actions = agent.get("action_distribution", {})
            top = sorted(actions.items(), key=lambda x: x[1], reverse=True)[:3]
            top_str = ", ".join(f"{k}: {v}" for k, v in top)
            beh_table.add_row(
                agent.get("agent_name", "N/A"),
                str(agent.get("notes_written", 0)),
                str(agent.get("ratings_given", 0)),
                str(agent.get("turn_count", 0)),
                agent.get("state", "N/A"),
                top_str,
            )
        console.print(beh_table)

    quality = attrs.get("note_quality", {})
    avg_score = quality.get("avg_helpfulness_score")
    avg_display = f"{avg_score}" if avg_score is not None else "N/A"
    quality_content = f"[bold]Avg helpfulness score:[/bold] {avg_display}"
    console.print(Panel(quality_content, title="[bold]Note Quality[/bold]"))

    quality_status = quality.get("notes_by_status", {})
    if quality_status:
        qs_table = Table(show_header=True, header_style="bold")
        qs_table.add_column("Status")
        qs_table.add_column("Count", justify="right")
        for status, count in sorted(quality_status.items()):
            qs_table.add_row(status, str(count))
        console.print(qs_table)

    quality_class = quality.get("notes_by_classification", {})
    if quality_class:
        qc_table = Table(show_header=True, header_style="bold")
        qc_table.add_column("Classification")
        qc_table.add_column("Count", justify="right")
        for cls, count in sorted(quality_class.items()):
            qc_table.add_row(cls, str(count))
        console.print(qc_table)


def _render_analysis_markdown(simulation_id: str, attrs: dict[str, Any], use_huuid: bool) -> None:
    lines: list[str] = []
    lines.append(f"# Simulation Analysis: {format_id(simulation_id, use_huuid)}")
    lines.append("")

    rating_dist = attrs.get("rating_distribution", {})
    overall = rating_dist.get("overall", {})
    total_ratings = rating_dist.get("total_ratings", 0)

    lines.append("## Rating Distribution")
    lines.append("")
    lines.append("| Rating | Count | % |")
    lines.append("|--------|-------|---|")
    for rating, count in sorted(overall.items()):
        pct = (count / total_ratings * 100) if total_ratings > 0 else 0
        lines.append(f"| {rating} | {count} | {pct:.1f}% |")
    total_pct = "100.0%" if total_ratings > 0 else "N/A"
    lines.append(f"| **Total** | **{total_ratings}** | **{total_pct}** |")
    lines.append("")

    per_agent = rating_dist.get("per_agent", [])
    if per_agent:
        lines.append("### Per-Agent Breakdown")
        lines.append("")
        for agent in per_agent:
            name = agent.get("agent_name", "N/A")
            total = agent.get("total", 0)
            lines.append(f"**{name}** ({total} ratings)")
            dist = agent.get("distribution", {})
            for r, c in sorted(dist.items()):
                lines.append(f"- {r}: {c}")
            lines.append("")

    consensus = attrs.get("consensus_metrics", {})
    lines.append("## Consensus Metrics")
    lines.append("")
    lines.append(f"- Mean agreement: {consensus.get('mean_agreement', 'N/A')}")
    lines.append(f"- Polarization index: {consensus.get('polarization_index', 'N/A')}")
    lines.append(f"- Notes with consensus: {consensus.get('notes_with_consensus', 'N/A')}")
    lines.append(
        f"- Notes with disagreement: {consensus.get('notes_with_disagreement', 'N/A')}"
    )
    lines.append(f"- Total notes rated: {consensus.get('total_notes_rated', 'N/A')}")
    lines.append("")

    scoring = attrs.get("scoring_coverage", {})
    lines.append("## Scoring Coverage")
    lines.append("")
    lines.append(f"- Current tier: {scoring.get('current_tier', 'N/A')}")
    lines.append(f"- Total scores computed: {scoring.get('total_scores_computed', 'N/A')}")
    lines.append("")

    tier_dist = scoring.get("tier_distribution", {})
    if tier_dist:
        lines.append("### Tier Distribution")
        lines.append("")
        lines.append("| Tier | Count |")
        lines.append("|------|-------|")
        for tier, count in sorted(tier_dist.items()):
            lines.append(f"| {tier} | {count} |")
        lines.append("")

    scorer_bd = scoring.get("scorer_breakdown", {})
    if scorer_bd:
        lines.append("### Scorer Breakdown")
        lines.append("")
        lines.append("| Scorer | Count |")
        lines.append("|--------|-------|")
        for scorer, count in sorted(scorer_bd.items()):
            lines.append(f"| {scorer} | {count} |")
        lines.append("")

    scoring_notes = scoring.get("notes_by_status", {})
    if scoring_notes:
        lines.append("### Notes by Status")
        lines.append("")
        lines.append("| Status | Count |")
        lines.append("|--------|-------|")
        for status, count in sorted(scoring_notes.items()):
            lines.append(f"| {status} | {count} |")
        lines.append("")

    behaviors = attrs.get("agent_behaviors", [])
    if behaviors:
        lines.append("## Agent Behaviors")
        lines.append("")
        lines.append("| Agent | Notes Written | Ratings Given | Turns | State |")
        lines.append("|-------|---------------|---------------|-------|-------|")
        for agent in behaviors:
            lines.append(
                f"| {agent.get('agent_name', 'N/A')}"
                f" | {agent.get('notes_written', 0)}"
                f" | {agent.get('ratings_given', 0)}"
                f" | {agent.get('turn_count', 0)}"
                f" | {agent.get('state', 'N/A')} |"
            )
        lines.append("")

    quality = attrs.get("note_quality", {})
    avg_score = quality.get("avg_helpfulness_score")
    avg_display = str(avg_score) if avg_score is not None else "N/A"

    lines.append("## Note Quality")
    lines.append("")
    lines.append(f"- Avg helpfulness score: {avg_display}")
    lines.append("")

    quality_status = quality.get("notes_by_status", {})
    if quality_status:
        lines.append("### Notes by Status")
        lines.append("")
        lines.append("| Status | Count |")
        lines.append("|--------|-------|")
        for status, count in sorted(quality_status.items()):
            lines.append(f"| {status} | {count} |")
        lines.append("")

    quality_class = quality.get("notes_by_classification", {})
    if quality_class:
        lines.append("### Notes by Classification")
        lines.append("")
        lines.append("| Classification | Count |")
        lines.append("|----------------|-------|")
        for cls, count in sorted(quality_class.items()):
            lines.append(f"| {cls} | {count} |")
        lines.append("")

    click.echo("\n".join(lines))


def _fetch_detailed_pages(
    client: Any,
    base_url: str,
    headers: dict[str, str],
    simulation_id: str,
) -> dict[str, list[dict[str, Any]]]:
    all_notes: list[dict[str, Any]] = []
    all_ratings: list[dict[str, Any]] = []
    all_requests: list[dict[str, Any]] = []
    all_agents: list[dict[str, Any]] = []
    page_number = 1

    with error_console.status("Fetching detailed analysis...") as status:
        while True:
            status.update(f"Fetching page {page_number} ({len(all_notes)} notes so far)")
            params = {"page[number]": page_number, "page[size]": 50}
            response = client.get(
                f"{base_url}/api/v2/simulations/{simulation_id}/analysis/detailed",
                headers=headers,
                params=params,
            )
            handle_jsonapi_error(response)
            result = response.json()

            for resource in result.get("data", []):
                attrs = resource.get("attributes", {})
                note = {
                    "note_id": attrs.get("note_id", ""),
                    "summary": attrs.get("summary", ""),
                    "classification": attrs.get("classification", ""),
                    "status": attrs.get("status", ""),
                    "helpfulness_score": attrs.get("helpfulness_score"),
                    "author_agent": attrs.get("author_agent_name", ""),
                    "request_id": attrs.get("request_id") or "",
                    "created_at": attrs.get("created_at"),
                }
                all_notes.append(note)
                for r in attrs.get("ratings", []):
                    all_ratings.append({
                        "note_id": attrs.get("note_id", ""),
                        "note_summary": attrs.get("summary", ""),
                        "rater_agent": r.get("rater_agent_name", ""),
                        "helpfulness_level": r.get("helpfulness_level", ""),
                        "created_at": r.get("created_at"),
                    })

            if page_number == 1:
                meta = result.get("meta", {})
                variance = meta.get("request_variance", {})
                all_requests.extend(variance.get("requests", []))
                all_agents.extend(meta.get("agents", []))

            links = result.get("links", {})
            if not links.get("next"):
                break
            page_number += 1

    error_console.print(f"Fetched {len(all_notes)} notes across {page_number} page(s)")
    return {"notes": all_notes, "ratings": all_ratings, "requests": all_requests, "agents": all_agents}


def _render_detailed_terminal(
    simulation_id: str, data: dict[str, list[dict[str, Any]]], use_huuid: bool
) -> None:
    notes = data.get("notes", [])
    ratings = data.get("ratings", [])
    requests = data.get("requests", [])
    agents = data.get("agents", [])

    console.print(
        Panel(
            f"[bold]Simulation:[/bold] {format_id(simulation_id, use_huuid)}\n"
            f"[bold]Notes:[/bold] {len(notes)}  "
            f"[bold]Ratings:[/bold] {len(ratings)}  "
            f"[bold]Requests:[/bold] {len(requests)}  "
            f"[bold]Agents:[/bold] {len(agents)}",
            title="[bold]Detailed Simulation Analysis[/bold]",
        )
    )

    if notes:
        note_table = Table(title="Notes", show_header=True, header_style="bold")
        note_table.add_column("Note ID")
        note_table.add_column("Summary", max_width=50)
        note_table.add_column("Classification")
        note_table.add_column("Status")
        note_table.add_column("Score", justify="right")
        note_table.add_column("Author")
        note_table.add_column("Request ID")
        note_table.add_column("Created At", width=20)
        for n in notes:
            score = n.get("helpfulness_score")
            score_str = f"{score}" if score is not None else "N/A"
            note_table.add_row(
                format_id(n.get("note_id", "N/A"), use_huuid),
                (n.get("summary", "") or "")[:50],
                n.get("classification", "N/A"),
                n.get("status", "N/A"),
                score_str,
                n.get("author_agent", "N/A"),
                format_id(n.get("request_id", "N/A"), use_huuid),
                (n.get("created_at", "") or "")[:19],
            )
        console.print(note_table)

    if ratings:
        rating_table = Table(title="Ratings", show_header=True, header_style="bold")
        rating_table.add_column("Note ID")
        rating_table.add_column("Note Summary", max_width=50)
        rating_table.add_column("Rater")
        rating_table.add_column("Helpfulness")
        rating_table.add_column("Created At", width=20)
        for r in ratings:
            rating_table.add_row(
                format_id(r.get("note_id", "N/A"), use_huuid),
                (r.get("note_summary", "") or "")[:50],
                r.get("rater_agent", "N/A"),
                r.get("helpfulness_level", "N/A"),
                (r.get("created_at", "") or "")[:19],
            )
        console.print(rating_table)

    if requests:
        req_table = Table(title="Requests", show_header=True, header_style="bold")
        req_table.add_column("Request ID")
        req_table.add_column("Content", max_width=60)
        req_table.add_column("Type")
        req_table.add_column("Notes", justify="right")
        req_table.add_column("Variance", justify="right")
        for req in requests:
            variance = req.get("variance_score")
            variance_str = f"{variance}" if variance is not None else "N/A"
            req_table.add_row(
                format_id(req.get("request_id", "N/A"), use_huuid),
                (req.get("content", "") or "")[:60],
                req.get("content_type", "N/A"),
                str(req.get("note_count", 0)),
                variance_str,
            )
        console.print(req_table)

    sorted_requests = sorted(
        [r for r in requests if r.get("variance_score") is not None],
        key=lambda r: r["variance_score"],
        reverse=True,
    )
    if sorted_requests:
        var_table = Table(
            title="Request Variance Summary (Most Varied First)",
            show_header=True,
            header_style="bold",
        )
        var_table.add_column("Request ID")
        var_table.add_column("Variance Score", justify="right")
        var_table.add_column("Content", max_width=60)
        var_table.add_column("Notes", justify="right")
        for req in sorted_requests:
            var_table.add_row(
                format_id(req.get("request_id", "N/A"), use_huuid),
                f"{req['variance_score']}",
                (req.get("content", "") or "")[:60],
                str(req.get("note_count", 0)),
            )
        console.print(var_table)

    if agents:
        agent_table = Table(title="Agents", show_header=True, header_style="bold")
        agent_table.add_column("Agent Name")
        agent_table.add_column("Model")
        agent_table.add_column("Memory Strategy")
        agent_table.add_column("Turn Count", justify="right")
        agent_table.add_column("State")
        agent_table.add_column("Token Count", justify="right")
        agent_table.add_column("Recent Actions", max_width=40)
        for a in agents:
            recent = ", ".join(str(x) for x in (a.get("recent_actions", []) or [])[:5])
            agent_table.add_row(
                a.get("agent_name", "N/A"),
                a.get("model_name", "N/A"),
                a.get("memory_compaction_strategy", "N/A"),
                str(a.get("turn_count", 0)),
                a.get("state", "N/A"),
                str(a.get("token_count", 0)),
                recent or "N/A",
            )
        console.print(agent_table)


def _escape_md(text: str) -> str:
    return text.replace("|", "\\|")


def _render_detailed_markdown(
    simulation_id: str, data: dict[str, list[dict[str, Any]]], use_huuid: bool
) -> None:
    notes = data.get("notes", [])
    ratings = data.get("ratings", [])
    requests = data.get("requests", [])
    agents = data.get("agents", [])

    lines: list[str] = []
    lines.append(f"# Detailed Simulation Analysis: {format_id(simulation_id, use_huuid)}")
    lines.append("")
    lines.append(f"- Notes: {len(notes)}")
    lines.append(f"- Ratings: {len(ratings)}")
    lines.append(f"- Requests: {len(requests)}")
    lines.append(f"- Agents: {len(agents)}")
    lines.append("")

    lines.append("## Notes")
    lines.append("")
    if notes:
        lines.append("| Note ID | Summary | Classification | Status | Score | Author | Request ID | Created At |")
        lines.append("|---------|---------|----------------|--------|-------|--------|------------|------------|")
        for n in notes:
            score = n.get("helpfulness_score")
            score_str = str(score) if score is not None else "N/A"
            summary = _escape_md((n.get("summary", "") or "")[:50])
            lines.append(
                f"| {format_id(n.get('note_id', 'N/A'), use_huuid)}"
                f" | {summary}"
                f" | {_escape_md(n.get('classification', 'N/A'))}"
                f" | {_escape_md(n.get('status', 'N/A'))}"
                f" | {score_str}"
                f" | {_escape_md(n.get('author_agent', 'N/A'))}"
                f" | {format_id(n.get('request_id', 'N/A'), use_huuid)}"
                f" | {(n.get('created_at', '') or '')[:19]} |"
            )
        lines.append("")
    else:
        lines.append("No notes found.")
        lines.append("")

    lines.append("## Ratings")
    lines.append("")
    if ratings:
        lines.append("| Note ID | Note Summary | Rater | Helpfulness | Created At |")
        lines.append("|---------|--------------|-------|-------------|------------|")
        for r in ratings:
            summary = _escape_md((r.get("note_summary", "") or "")[:50])
            lines.append(
                f"| {format_id(r.get('note_id', 'N/A'), use_huuid)}"
                f" | {summary}"
                f" | {_escape_md(r.get('rater_agent', 'N/A'))}"
                f" | {_escape_md(r.get('helpfulness_level', 'N/A'))}"
                f" | {(r.get('created_at', '') or '')[:19]} |"
            )
        lines.append("")
    else:
        lines.append("No ratings found.")
        lines.append("")

    lines.append("## Requests")
    lines.append("")
    if requests:
        lines.append("| Request ID | Content | Type | Notes | Variance Score |")
        lines.append("|------------|---------|------|-------|----------------|")
        for req in requests:
            variance = req.get("variance_score")
            variance_str = str(variance) if variance is not None else "N/A"
            content = _escape_md((req.get("content", "") or "")[:60])
            lines.append(
                f"| {format_id(req.get('request_id', 'N/A'), use_huuid)}"
                f" | {content}"
                f" | {_escape_md(req.get('content_type', 'N/A'))}"
                f" | {req.get('note_count', 0)}"
                f" | {variance_str} |"
            )
        lines.append("")
    else:
        lines.append("No requests found.")
        lines.append("")

    sorted_requests = sorted(
        [r for r in requests if r.get("variance_score") is not None],
        key=lambda r: r["variance_score"],
        reverse=True,
    )
    if sorted_requests:
        lines.append("## Request Variance Summary")
        lines.append("")
        lines.append("Requests ranked by variance score (most varied first):")
        lines.append("")
        lines.append("| Rank | Request ID | Variance Score | Content | Notes |")
        lines.append("|------|------------|----------------|---------|-------|")
        for rank, req in enumerate(sorted_requests, 1):
            content = _escape_md((req.get("content", "") or "")[:60])
            lines.append(
                f"| {rank}"
                f" | {format_id(req.get('request_id', 'N/A'), use_huuid)}"
                f" | {req['variance_score']}"
                f" | {content}"
                f" | {req.get('note_count', 0)} |"
            )
        lines.append("")

    lines.append("## Agents")
    lines.append("")
    if agents:
        lines.append("| Agent Name | Personality | Model | Memory Strategy | Turn Count | State | Token Count | Recent Actions | Last Messages |")
        lines.append("|------------|-------------|-------|-----------------|------------|-------|-------------|----------------|---------------|")
        for a in agents:
            personality = _escape_md((a.get("personality", "") or "")[:80])
            recent = _escape_md(", ".join(str(x) for x in (a.get("recent_actions", []) or [])[:5]))
            msgs = a.get("last_messages", []) or []
            msg_summary = _escape_md(f"{len(msgs)} messages" if msgs else "None")
            lines.append(
                f"| {_escape_md(a.get('agent_name', 'N/A'))}"
                f" | {personality}"
                f" | {_escape_md(a.get('model_name', 'N/A'))}"
                f" | {_escape_md(a.get('memory_compaction_strategy', 'N/A'))}"
                f" | {a.get('turn_count', 0)}"
                f" | {_escape_md(a.get('state', 'N/A'))}"
                f" | {a.get('token_count', 0)}"
                f" | {recent or 'N/A'}"
                f" | {msg_summary} |"
            )
        lines.append("")
    else:
        lines.append("No agent data found.")
        lines.append("")

    click.echo("\n".join(lines))


_XLSX_MAX_CELL_LENGTH = 32767
_XLSX_TRUNCATION_SUFFIX = "...[truncated]"


def _truncate_for_xlsx(value: object) -> object:
    if isinstance(value, str) and len(value) > _XLSX_MAX_CELL_LENGTH:
        return value[: _XLSX_MAX_CELL_LENGTH - len(_XLSX_TRUNCATION_SUFFIX)] + _XLSX_TRUNCATION_SUFFIX
    return value


_PART_KIND_TO_ROLE = {
    "user-prompt": "user",
    "text": "assistant",
    "tool-call": "tool-call",
    "tool-return": "tool-return",
    "retry-prompt": "retry",
}


def _format_pydantic_ai_messages(msgs: list[dict[str, Any]]) -> str:
    lines: list[str] = []
    for msg in msgs:
        parts = msg.get("parts") or []
        for part in parts:
            part_kind = part.get("part_kind", "")
            if part_kind == "system-prompt":
                continue
            role = _PART_KIND_TO_ROLE.get(part_kind, part_kind)
            if part_kind == "tool-call":
                raw = part.get("tool_name", "")
            else:
                raw = part.get("content", "")
            if isinstance(raw, str):
                content = raw
            else:
                content = json.dumps(raw, default=str)
            lines.append(f"{role}: {content}")
    return "\n".join(lines)


def _render_detailed_xlsx(
    simulation_id: str,
    data: dict[str, list[dict[str, Any]]],
    output_path: str | None,
    use_huuid: bool = False,
) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        error_console.print("[red]Error:[/red] openpyxl is required for xlsx export. Install with: uv add openpyxl")
        sys.exit(1)

    notes = data.get("notes", [])
    ratings = data.get("ratings", [])
    requests = data.get("requests", [])

    wb = Workbook()

    ws_notes = wb.active
    assert ws_notes is not None
    ws_notes.title = "Notes"
    note_headers = ["Note ID", "Summary", "Classification", "Status", "Helpfulness Score", "Author Agent", "Request ID", "Created At"]
    ws_notes.append(note_headers)
    for n in notes:
        ws_notes.append([
            format_id(n.get("note_id"), use_huuid),
            n.get("summary", ""),
            n.get("classification", ""),
            n.get("status", ""),
            n.get("helpfulness_score"),
            n.get("author_agent", ""),
            format_id(n.get("request_id"), use_huuid),
            n.get("created_at", ""),
        ])

    ws_ratings = wb.create_sheet("Ratings")
    rating_headers = ["Note ID", "Note Summary", "Rater Agent", "Helpfulness Level", "Created At"]
    ws_ratings.append(rating_headers)
    for r in ratings:
        ws_ratings.append([
            format_id(r.get("note_id"), use_huuid),
            r.get("note_summary", "") or "",
            r.get("rater_agent", ""),
            r.get("helpfulness_level", ""),
            r.get("created_at", ""),
        ])

    ws_requests = wb.create_sheet("Requests")
    request_headers = ["Request ID", "Content", "Content Type", "Note Count", "Variance Score"]
    ws_requests.append(request_headers)
    for req in requests:
        ws_requests.append([
            format_id(req.get("request_id"), use_huuid),
            req.get("content", ""),
            req.get("content_type", ""),
            req.get("note_count", 0),
            req.get("variance_score"),
        ])

    ws_agents = wb.create_sheet("Agents")
    agent_headers = [
        "Agent Name", "Personality Prompt", "Model",
        "Memory Compaction Strategy", "Turn Count", "State",
        "Token Count", "Recent Actions", "Last 30 Messages",
    ]
    ws_agents.append(agent_headers)
    agents = data.get("agents", [])
    for a in agents:
        recent = ", ".join(str(x) for x in (a.get("recent_actions", []) or []))
        msgs = a.get("last_messages", []) or []
        msg_text = _format_pydantic_ai_messages(msgs) if msgs else ""
        ws_agents.append([
            a.get("agent_name", ""),
            a.get("personality", ""),
            a.get("model_name", ""),
            a.get("memory_compaction_strategy", ""),
            a.get("turn_count", 0),
            a.get("state", ""),
            a.get("token_count", 0),
            recent,
            msg_text,
        ])

    for ws in [ws_notes, ws_ratings, ws_requests, ws_agents]:
        for row in ws.iter_rows():
            for cell in row:
                cell.value = _truncate_for_xlsx(cell.value)

    for ws in [ws_notes, ws_ratings, ws_requests, ws_agents]:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value or ""))
                    if cell_len > max_length:
                        max_length = cell_len
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[col_letter].width = adjusted_width

    from openpyxl.styles import Alignment, Font

    header_font = Font(name="IBM Plex Sans Condensed", bold=True)
    default_font = Font(name="IBM Plex Sans Condensed")
    default_alignment = Alignment(vertical="top")
    for ws in [ws_notes, ws_ratings, ws_requests, ws_agents]:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = header_font if cell.row == 1 else default_font
                cell.alignment = default_alignment

    if not output_path:
        output_path = f"simulation-{simulation_id}-detailed.xlsx"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    console.print(f"[green]\u2713[/green] Saved detailed analysis to [bold]{output_path}[/bold]")
