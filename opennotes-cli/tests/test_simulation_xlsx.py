from __future__ import annotations

import os
from unittest.mock import MagicMock, patch

import pytest
from click.testing import CliRunner

from opennotes_cli.cli import cli


@pytest.fixture()
def runner() -> CliRunner:
    return CliRunner()


def _make_csrf_response() -> MagicMock:
    resp = MagicMock()
    resp.status_code = 200
    return resp


def _make_detailed_response(simulation_id: str) -> MagicMock:
    body = {
        "data": [
            {
                "type": "simulation-detailed-notes",
                "id": "note-001",
                "attributes": {
                    "note_id": "note-001",
                    "summary": "Claim is misleading due to cherry-picked statistics",
                    "classification": "MISINFORMATION",
                    "status": "scored",
                    "helpfulness_score": 0.85,
                    "author_agent_name": "Skeptic",
                    "author_agent_instance_id": "skeptic-1",
                    "request_id": "req-001",
                    "created_at": "2026-03-01T10:00:00Z",
                    "ratings": [
                        {
                            "rater_agent_name": "Optimist",
                            "rater_agent_instance_id": "optimist-1",
                            "helpfulness_level": "HELPFUL",
                            "created_at": "2026-03-01T10:10:00Z",
                        }
                    ],
                },
            },
            {
                "type": "simulation-detailed-notes",
                "id": "note-002",
                "attributes": {
                    "note_id": "note-002",
                    "summary": "Source is reliable and well-cited",
                    "classification": "ACCURATE",
                    "status": "scored",
                    "helpfulness_score": 0.72,
                    "author_agent_name": "Optimist",
                    "author_agent_instance_id": "optimist-1",
                    "request_id": "req-002",
                    "created_at": "2026-03-01T10:05:00Z",
                    "ratings": [
                        {
                            "rater_agent_name": "Skeptic",
                            "rater_agent_instance_id": "skeptic-1",
                            "helpfulness_level": "SOMEWHAT_HELPFUL",
                            "created_at": "2026-03-01T10:15:00Z",
                        }
                    ],
                },
            },
        ],
        "links": {"self": f"/api/v2/simulations/{simulation_id}/analysis/detailed", "next": None},
        "meta": {
            "count": 2,
            "request_variance": {
                "requests": [
                    {
                        "request_id": "req-001",
                        "content": "https://example.com/article-1",
                        "content_type": "url",
                        "note_count": 3,
                        "variance_score": 0.92,
                    },
                    {
                        "request_id": "req-002",
                        "content": "Some user-submitted text",
                        "content_type": "text",
                        "note_count": 2,
                        "variance_score": 0.45,
                    },
                ],
                "total_requests": 2,
            },
            "agents": [
                {
                    "agent_instance_id": "inst-001",
                    "agent_name": "Skeptic",
                    "personality": "A skeptical fact-checker who questions claims",
                    "model_name": "openai:gpt-4o-mini",
                    "memory_compaction_strategy": "sliding_window",
                    "turn_count": 5,
                    "state": "active",
                    "token_count": 1200,
                    "recent_actions": ["write_note", "rate_note", "write_note"],
                    "last_messages": [
                        {
                            "parts": [
                                {"content": "Review this claim", "part_kind": "user-prompt", "timestamp": "2026-03-01T10:00:00Z"},
                            ],
                            "kind": "request",
                        },
                        {
                            "parts": [
                                {"content": "I'll check the sources", "part_kind": "text"},
                            ],
                            "kind": "response",
                            "model_name": "openai:gpt-4o-mini",
                            "timestamp": "2026-03-01T10:00:01Z",
                        },
                    ],
                },
                {
                    "agent_instance_id": "inst-002",
                    "agent_name": "Optimist",
                    "personality": "An optimistic analyst who sees the good in content",
                    "model_name": "openai:gpt-4o-mini",
                    "memory_compaction_strategy": "summarize_and_prune",
                    "turn_count": 3,
                    "state": "active",
                    "token_count": 800,
                    "recent_actions": ["rate_note"],
                    "last_messages": [],
                },
            ],
        },
    }
    resp = MagicMock()
    resp.status_code = 200
    resp.json.return_value = body
    return resp


SIM_ID = "019536b8-bdb2-7c81-8975-77f5c3dbdff8"


def _make_xlsx_client() -> MagicMock:
    csrf_resp = _make_csrf_response()
    detail_resp = _make_detailed_response(SIM_ID)
    mock_client = MagicMock()
    mock_client.get.side_effect = [csrf_resp, detail_resp]
    mock_client.cookies = MagicMock()
    mock_client.cookies.get.return_value = "test-csrf"
    return mock_client


HUUID_SIM_ID = "Vudrotlab-Kuvkattor-Tevzelpim-Liksiksas"


class TestXlsxHuuidInput:
    def test_xlsx_accepts_huuid(self, runner: CliRunner) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test-huuid.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                result = runner.invoke(
                    cli,
                    [
                        "--local",
                        "simulation",
                        "analysis",
                        "--detailed",
                        "--format",
                        "xlsx",
                        "--output",
                        output_file,
                        HUUID_SIM_ID,
                    ],
                )
            assert result.exit_code == 0, f"CLI failed: {result.output}"
            assert os.path.exists(output_file)


class TestXlsxRequiresDetailed:
    def test_xlsx_without_detailed_errors(self, runner: CliRunner) -> None:
        result = runner.invoke(
            cli, ["--local", "simulation", "analysis", "--format", "xlsx", SIM_ID]
        )
        assert result.exit_code != 0
        assert "requires --detailed" in result.output or "requires --detailed" in (result.output + str(result.exception or ""))


class TestXlsxOutput:
    def test_xlsx_creates_file(self, runner: CliRunner, tmp_path: object) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test-output.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                result = runner.invoke(
                    cli,
                    [
                        "--local",
                        "simulation",
                        "analysis",
                        "--detailed",
                        "--format",
                        "xlsx",
                        "--output",
                        output_file,
                        SIM_ID,
                    ],
                )
            assert result.exit_code == 0, f"CLI failed: {result.output}"
            assert os.path.exists(output_file)

    def test_xlsx_default_filename(self, runner: CliRunner) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                old_cwd = os.getcwd()
                try:
                    os.chdir(tmpdir)
                    result = runner.invoke(
                        cli,
                        [
                            "--local",
                            "simulation",
                            "analysis",
                            "--detailed",
                            "--format",
                            "xlsx",
                            SIM_ID,
                        ],
                    )
                finally:
                    os.chdir(old_cwd)
            assert result.exit_code == 0
            expected_file = os.path.join(tmpdir, f"simulation-{SIM_ID}-detailed.xlsx")
            assert os.path.exists(expected_file)


class TestXlsxNotesSheet:
    def test_notes_sheet_has_correct_headers(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Notes"]
            headers = [cell.value for cell in ws[1]]
            assert "Note ID" in headers
            assert "Summary" in headers
            assert "Classification" in headers
            assert "Status" in headers
            assert "Helpfulness Score" in headers
            assert "Author Agent" in headers
            assert "Request ID" in headers
            assert "Created At" in headers

    def test_notes_sheet_has_data_rows(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Notes"]
            assert ws.max_row == 3
            assert ws.cell(row=2, column=1).value.replace("\u200b", "") == "note-001"
            assert ws.cell(row=3, column=1).value.replace("\u200b", "") == "note-002"


class TestXlsxRatingsSheet:
    def test_ratings_sheet_has_correct_headers(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Ratings"]
            headers = [cell.value for cell in ws[1]]
            assert "Note ID" in headers
            assert "Note Summary" in headers
            assert "Rater Agent" in headers
            assert "Helpfulness Level" in headers
            assert "Created At" in headers

    def test_ratings_sheet_preserves_full_summary(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Ratings"]
            summary_val = ws.cell(row=2, column=2).value
            assert summary_val == "Claim is misleading due to cherry-picked statistics"


class TestXlsxRequestsSheet:
    def test_requests_sheet_has_correct_headers(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Requests"]
            headers = [cell.value for cell in ws[1]]
            assert "Request ID" in headers
            assert "Content" in headers
            assert "Content Type" in headers
            assert "Note Count" in headers
            assert "Variance Score" in headers

    def test_requests_sheet_has_data(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Requests"]
            assert ws.max_row == 3
            assert ws.cell(row=2, column=1).value.replace("\u200b", "") == "req-001"
            assert ws.cell(row=2, column=4).value == 3
            assert ws.cell(row=2, column=5).value == 0.92


class TestXlsxOpensCorrectly:
    def test_workbook_has_four_sheets(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            assert wb.sheetnames == ["Notes", "Ratings", "Requests", "Agents"]

    def test_columns_have_width(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Notes"]
            a_width = ws.column_dimensions["A"].width
            assert a_width > 0


class TestXlsxAgentsSheet:
    def test_agents_sheet_has_correct_headers(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Agents"]
            headers = [cell.value for cell in ws[1]]
            assert "Agent Name" in headers
            assert "Personality Prompt" in headers
            assert "Model" in headers
            assert "Memory Compaction Strategy" in headers
            assert "Turn Count" in headers
            assert "State" in headers
            assert "Token Count" in headers
            assert "Recent Actions" in headers
            assert "Last 30 Messages" in headers

    def test_agents_sheet_has_data(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Agents"]
            assert ws.max_row == 3
            assert ws.cell(row=2, column=1).value == "Skeptic"
            assert ws.cell(row=3, column=1).value == "Optimist"

    def test_agents_sheet_messages_formatted(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Agents"]
            msg_col = 9
            msg_text = ws.cell(row=2, column=msg_col).value
            assert "user: Review this claim" in msg_text
            assert "assistant: I'll check the sources" in msg_text
            empty_msg = ws.cell(row=3, column=msg_col).value
            assert empty_msg == "" or empty_msg is None

    def test_agents_sheet_multipart_messages(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        body = _make_detailed_response(SIM_ID).json.return_value
        body["meta"]["agents"][0]["last_messages"] = [
            {
                "parts": [
                    {"content": "System prompt", "part_kind": "system-prompt"},
                    {"content": "User question", "part_kind": "user-prompt", "timestamp": "2026-03-01T10:00:00Z"},
                ],
                "kind": "request",
            },
            {
                "parts": [
                    {"content": "Let me check", "part_kind": "text"},
                    {"tool_name": "search", "args": {"q": "test"}, "tool_call_id": "tc1", "part_kind": "tool-call"},
                ],
                "kind": "response",
                "model_name": "openai:gpt-4o-mini",
                "timestamp": "2026-03-01T10:00:01Z",
            },
            {
                "parts": [
                    {"tool_name": "search", "content": "results", "tool_call_id": "tc1", "part_kind": "tool-return"},
                ],
                "kind": "request",
            },
        ]
        resp = MagicMock()
        resp.status_code = 200
        resp.json.return_value = body

        csrf_resp = _make_csrf_response()
        mock_client = MagicMock()
        mock_client.get.side_effect = [csrf_resp, resp]
        mock_client.cookies = MagicMock()
        mock_client.cookies.get.return_value = "test-csrf"

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Agents"]
            msg_text = ws.cell(row=2, column=9).value
            assert "user: User question" in msg_text
            assert "system-prompt" not in msg_text.lower() or "System prompt" not in msg_text
            assert "assistant: Let me check" in msg_text
            assert "tool-call: search" in msg_text
            assert "tool-return: results" in msg_text

    def test_agents_sheet_personality_not_truncated(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Agents"]
            assert ws.cell(row=2, column=2).value == "A skeptical fact-checker who questions claims"


class TestXlsxFontAndAlignment:
    def test_cells_use_ibm_plex_sans_condensed(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Notes"]
            assert ws.cell(row=1, column=1).font.name == "IBM Plex Sans Condensed"
            assert ws.cell(row=2, column=1).font.name == "IBM Plex Sans Condensed"

    def test_cells_use_top_alignment(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            ws = wb["Notes"]
            assert ws.cell(row=1, column=1).alignment.vertical == "top"
            assert ws.cell(row=2, column=1).alignment.vertical == "top"

    def test_font_applied_to_all_sheets(self, runner: CliRunner) -> None:
        import tempfile

        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = os.path.join(tmpdir, "test.xlsx")
            mock_client = _make_xlsx_client()
            with patch("opennotes_cli.cli.httpx.Client", return_value=mock_client):
                runner.invoke(
                    cli,
                    ["--local", "simulation", "analysis", "--detailed", "--format", "xlsx", "--output", output_file, SIM_ID],
                )
            wb = load_workbook(output_file)
            for sheet_name in ["Notes", "Ratings", "Requests", "Agents"]:
                ws = wb[sheet_name]
                assert ws.cell(row=1, column=1).font.name == "IBM Plex Sans Condensed"


class TestFormatPydanticAiMessages:
    def test_parts_none_treated_as_empty(self) -> None:
        from opennotes_cli.commands.simulation import _format_pydantic_ai_messages

        msgs = [{"parts": None, "kind": "request"}]
        assert _format_pydantic_ai_messages(msgs) == ""

    def test_parts_key_missing_treated_as_empty(self) -> None:
        from opennotes_cli.commands.simulation import _format_pydantic_ai_messages

        msgs = [{"kind": "request"}]
        assert _format_pydantic_ai_messages(msgs) == ""

    def test_dict_content_serialized_as_json(self) -> None:
        from opennotes_cli.commands.simulation import _format_pydantic_ai_messages

        msgs = [
            {
                "parts": [
                    {
                        "tool_name": "search",
                        "content": {"results": [1, 2, 3]},
                        "tool_call_id": "tc1",
                        "part_kind": "tool-return",
                    }
                ],
                "kind": "request",
            }
        ]
        result = _format_pydantic_ai_messages(msgs)
        assert "tool-return:" in result
        assert '"results"' in result
        assert "[1, 2, 3]" in result

    def test_list_content_serialized_as_json(self) -> None:
        from opennotes_cli.commands.simulation import _format_pydantic_ai_messages

        msgs = [
            {
                "parts": [
                    {
                        "content": [{"type": "text", "text": "hello"}],
                        "part_kind": "user-prompt",
                        "timestamp": "2026-03-01T10:00:00Z",
                    }
                ],
                "kind": "request",
            }
        ]
        result = _format_pydantic_ai_messages(msgs)
        assert "user:" in result
        assert '"text"' in result
        assert '"hello"' in result

    def test_string_content_unchanged(self) -> None:
        from opennotes_cli.commands.simulation import _format_pydantic_ai_messages

        msgs = [
            {
                "parts": [
                    {"content": "plain text", "part_kind": "text"},
                ],
                "kind": "response",
            }
        ]
        assert _format_pydantic_ai_messages(msgs) == "assistant: plain text"


class TestOutputFlag:
    def test_output_flag_in_help(self, runner: CliRunner) -> None:
        result = runner.invoke(cli, ["simulation", "analysis", "--help"])
        assert result.exit_code == 0
        assert "--output" in result.output
